import dataclasses
import glob
import json
import logging
import os
import traceback
from collections import defaultdict
from tqdm import tqdm

SAVE_DIR = "/mnt/nas03/ainno_robot_datasets/AInnoRobotDatasets/statistics"
UPDATE_FILE = os.path.join(SAVE_DIR, f"update.txt")


def load_latest_dataset_statistics():
    file_path = open(UPDATE_FILE).read().strip()
    return json.load(open(file_path))


def get_dataset_statistics(dataset, dataset_name=None):
    import numpy as np

    dataset_iter = iter(dataset)

    statistics = {
        "action": defaultdict(dict),
        "state": defaultdict(dict),
    }
    values = {
        "action": defaultdict(list),
        "state": defaultdict(list),
    }
    num_samples = 0
    import multiprocessing as mp

    process_info = mp.current_process().name.split("-")
    if len(process_info) == 1:
        task_id = 0
    else:
        task_id = process_info[-1]

    for step in tqdm(
        dataset_iter,
        desc=f"[{task_id:2}] statistic steps:",
        position=int(task_id),
        postfix=dataset_name,
        mininterval=1,
        leave=False,
    ):
        num_samples += 1
        # continue
        for k, v in step["actions"][0][0].items():
            if v is None:
                continue
            values["action"][k].append(v)
        for k, v in step["states"][0].items():
            if v is None:
                continue
            values["state"][k].append(v)

    for t in ["action", "state"]:
        for k, v in values[t].items():
            # 如果维度不一致，需要padding
            if isinstance(v[0], list):
                dim_list = [len(x) for x in v]
                if max(dim_list) != min(dim_list):
                    dim = max(dim_list)
                    for i in range(len(v)):
                        v[i] = v[i] + [np.nan] * (dim - len(v[i]))
            try:
                v = np.array(v).astype(float)
            except Exception as e:
                logging.error(
                    f"统计失败，错误信息：{e}, trace = {traceback.format_exc()}"
                )
                continue

            s = {
                "mean": np.nanmean(v, 0).tolist(),
                "std": np.nanstd(v, 0).tolist(),
                "max": np.nanmax(v, 0).tolist(),
                "min": np.nanmin(v, 0).tolist(),
                "p99": np.nanquantile(v, 0.99, 0).tolist(),
                "p01": np.nanquantile(v, 0.01, 0).tolist(),
            }
            for key, value in s.items():
                statistics[t][key][k] = value
    statistics["metadata"] = {
        "num_samples": num_samples,
    }
    return statistics


def calculate_dataset_statistics_by_episode_files(dataset_name, files):
    result = dict(choices=[])
    try:
        config = dict(
            observation_window_size=1,
            action_chunk_size=1,
            action_postpone=0,
            action_stride=1,
        )
        from dataloader2 import build_dataset

        dataset = build_dataset(
            episode_files=files,
            shuffle_episodes=False,
            load_image=False,
            **config,
        )
        s = get_dataset_statistics(dataset, dataset_name=dataset_name)
        s["metadata"]["num_episodes"] = len(files)
        result["choices"].append(dict(config=config, statistics=s))
    except Exception as e:
        logging.error(f"Dataset:{dataset_name},  traceback={traceback.format_exc()}")
        result["choices"].append(dict(config=dataclasses.asdict(config), error=f"{e}"))

    return result


def data_profile_for_tasks(
    dataset_info: dict[str, list[str]], output_path: str
) -> None:
    # 初始化统计数据结构
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    robot_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "total_files": 0,
                "total_steps": 0,
                "operators": defaultdict(lambda: {"files": 0, "steps": 0}),
            }
        )
    )
    robot_amount = defaultdict(lambda: list())

    # 使用tqdm显示进度
    total_files = sum(len(files) for files in dataset_info.values())
    with tqdm(total=total_files, desc="处理文件中") as pbar:
        for robot_name, files in dataset_info.items():
            for filename in files:
                try:
                    with open(filename, "r", encoding="utf-8") as f:
                        data = json.load(f)

                        task_name = data["metadata"]["task_name"]
                        operator = data["metadata"]["operator"]
                        num_steps = data["metadata"]["num_steps"]
                        experiment_time = data["metadata"]["experiment_time"]

                        # 更新总体统计
                        robot_stats[robot_name][task_name]["total_files"] += 1
                        robot_stats[robot_name][task_name]["total_steps"] += num_steps

                        # 更新操作员统计
                        robot_stats[robot_name][task_name]["operators"][operator][
                            "files"
                        ] += 1
                        robot_stats[robot_name][task_name]["operators"][operator][
                            "steps"
                        ] += num_steps

                        robot_amount[robot_name].append(
                            {"datetime": experiment_time, "count": num_steps}
                        )
                except Exception as e:
                    logging.info(f"处理文件 {filename} 时出错: {str(e)}")
                finally:
                    pbar.update(1)

    # 创建Excel写入器
    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    if not robot_stats:
        logging.info("警告：没有找到任何数据，将创建一个空的工作表")
        pd.DataFrame().to_excel(writer, index=False, sheet_name="Sheet1")

    # 为每个robot创建独立的sheet
    for robot_name, task_stats in robot_stats.items():
        # 准备数据用于DataFrame
        rows = []
        for task_name, stats in task_stats.items():
            # 添加任务总体统计行（作为该任务的第一行）
            rows.append(
                {
                    "序号": "",
                    "任务名称": task_name,
                    "操作员": "总体",
                    "文件数": stats["total_files"],
                    "总步数": stats["total_steps"],
                    "平均步数": (
                        round(stats["total_steps"] / stats["total_files"], 1)
                        if stats["total_files"] > 0
                        else 0
                    ),
                    "总时长(分钟)": round(stats["total_steps"] / 1800, 1),
                }
            )

            # 添加每个操作员的统计行
            for operator, op_stats in stats["operators"].items():
                rows.append(
                    {
                        "序号": "",
                        "任务名称": task_name,
                        "操作员": operator.replace("\u001bOP", "").replace(
                            "\u001bOS", ""
                        ),
                        "文件数": op_stats["files"],
                        "总步数": op_stats["steps"],
                        "平均步数": (
                            round(op_stats["steps"] / op_stats["files"], 1)
                            if op_stats["files"] > 0
                            else 0
                        ),
                        "总时长(分钟)": round(op_stats["steps"] / 1800, 1),
                    }
                )

        # 创建DataFrame并排序
        df = pd.DataFrame(rows)
        df = df.sort_values(["任务名称", "操作员"], ascending=[True, False])

        # 写入到json, 当excel写入出错时，可以方便查看原因
        df.to_json(
            os.path.splitext(output_path)[0] + f"_{robot_name}.json",
            orient="records",
            force_ascii=False,
            indent=4,
        )

        # 写入到对应的sheet
        df.to_excel(writer, index=False, sheet_name=robot_name)

        # 获取工作表
        worksheet = writer.sheets[robot_name]

        # 设置列宽
        column_widths = {
            "A": 8,  # 序号
            "B": 30,  # 任务名称
            "C": 15,  # 操作员
            "D": 10,  # 文件数
            "E": 10,  # 总步数
            "F": 10,  # 平均步数
            "G": 15,  # 总时长
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置标题行样式
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置数据行样式
        data_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 合并任务名称单元格并设置序号
        current_task = None
        start_row = 2
        task_rows = {}

        # 首先收集每个任务的行范围
        for row in range(2, len(df) + 2):
            task_name = worksheet.cell(row=row, column=2).value
            if task_name != current_task:
                if current_task is not None:
                    task_rows[current_task] = (start_row, row - 1)
                current_task = task_name
                start_row = row
        if current_task is not None:
            task_rows[current_task] = (start_row, len(df) + 1)

        # 合并单元格并设置序号
        for task_name, (start_row, end_row) in task_rows.items():
            # 合并任务名称单元格
            worksheet.merge_cells(f"B{start_row}:B{end_row}")
            # 合并序号单元格
            worksheet.merge_cells(f"A{start_row}:A{end_row}")
            # 设置序号
            worksheet.cell(row=start_row, column=1).value = (
                list(task_rows.keys()).index(task_name) + 1
            )

            # 设置"总体"行的字体加粗
            for col in ["C", "D", "E", "F", "G"]:
                cell = worksheet.cell(row=start_row, column=ord(col) - ord("A") + 1)
                cell.font = Font(bold=True)

        # 设置所有单元格的样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border

    # 记录总体统计数据
    for robot_name, stats in robot_amount.items():
        robot_amount_df = pd.DataFrame(stats)
        # robot_amount_df = pd.DataFrame.from_dict(stats, orient='index').reset_index()

        # 1. 将字符串转换为 datetime 类型
        robot_amount_df["datetime"] = pd.to_datetime(
            robot_amount_df["datetime"], format="%Y%m%d%H%M%S"
        )

        # 2. 按周聚合 (每周的总和，周一开始)
        weekly_df = (
            robot_amount_df.resample(
                "W-MON",
                on="datetime",
                closed="left",  # 包含周一，不包含下周一
                label="left",  # 用周一作为标签
            )
            .sum()
            .reset_index()
        )
        weekly_df.columns = ["week_start_date", "steps"]
        weekly_df["总时长(分钟)"] = (weekly_df["steps"] / 1800).round(1)

        # 3. 按月聚合 (每月的总和)
        monthly_df = (
            robot_amount_df.resample(
                "MS",
                on="datetime",
                closed="left",
                label="left",
            )
            .sum()
            .reset_index()
        )
        monthly_df.columns = ["month_start_date", "steps"]
        monthly_df["总时长(分钟)"] = (monthly_df["steps"] / 1800).round(1)

        # 4. 将数据写入 Excel
        weekly_df.to_excel(writer, index=False, sheet_name=f"{robot_name}周汇总")
        monthly_df.to_excel(writer, index=False, sheet_name=f"{robot_name}月汇总")

        # 5. 格式配置
        for sheet_name in [f"{robot_name}周汇总", f"{robot_name}月汇总"]:
            # 获取工作表
            worksheet = writer.sheets[sheet_name]

            # 设置列宽
            column_widths = {
                "A": 30,  # 日期
                "B": 10,  # 数据量
                "C": 15,  # 时长
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 设置标题行样式
            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    # 保存Excel文件
    writer.close()
